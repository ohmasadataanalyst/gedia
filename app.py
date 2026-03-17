import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import io
import re

st.set_page_config(page_title="Geidea & Foodics Summary Generator", layout="wide")

# Terminal ID → Branch Code (e.g. "B01", "B02", …)
TERMINAL_BRANCH_CODE_MAP = {
    "8189573201099250": "B01", "8189572201099250": "B01", "8189568401099240": "B01",
    "1551813501455580": "B02", "8189580801099270": "B02", "8189569601099240": "B02", "18680360": "B02",
    "1551824001455540": "B03", "8109111101030290": "B03", "63189123": "B03",
    "1551835401455550": "B04", "8189570001099240": "B04", "8109110701030290": "B04",
    "1551842301455540": "B05", "1551848901455540": "B05",
    "63189498": "B06", "63189502": "B06",
    "63189106": "B07", "63189110": "B07",
    "1554733501483710": "B08", "8189582401099270": "B08",
    "1551822701455590": "B09", "63189490": "B09", "63933957": "B09",
    "1551838501455590": "B10", "63189504": "B10",
    "1554861501483710": "B11", "8189567201099240": "B11", "63189494": "B11",
    "1551865401455580": "B12", "63189100": "B12", "63189101": "B12", "63189103": "B12",
    "8189582001099270": "B13", "63189105": "B13",
    "8189568001099240": "B14", "63188996": "B14",
    "63189121": "B15", "63189122": "B15", "63189124": "B15",
    "8189583201099280": "B16", "63189107": "B16",
    "63189493": "B17", "63189496": "B17",
    "63189508": "B18", "63189510": "B18", "63189512": "B18",
    "1551874401455580": "B19", "8189566401099230": "B19", "63189113": "B19",
    "8189578001099260": "B22", "8189574801099250": "B22",
    "63934017": "B21",
    "8189579201099270": "B26", "8189577201099260": "B26",
    "8189565201099230": "B27", "8189564801099230": "B27",
    "8189579601099270": "B28",
    "8189568801099240": "B30", "8189566801099230": "B30", "8189566001099230": "B30",
    "8189564001099230": "B31", "8189563601099230": "B31",
    "8189578401099260": "B33", "8189571201099250": "B33", "8189570801099240": "B33", "8189570401099240": "B33",
    "1551816501455580": "B34", "8189581201099270": "B34", "8189578801099260": "B34",
    "1551829301455570": "B35", "1551843501455570": "B35", "1551892301455570": "B35",
    "64729693": "B36", "64729694": "B36", "64729695": "B36", "64729696": "B36",
    "8144089701539140": "B37", "8199376901539140": "B37", "8134184401539140": "B37", "8145052001539140": "B37",
    "5567115001585440": "B39", "5567068001585440": "B39", "5566976601585440": "B39",
}

# Branch Code → Branch Name
BRANCH_CODE_NAME_MAP = {
    "B01": "NURUH", "B02": "KHRUH", "B03": "GHRUH", "B04": "NSRUH", "B05": "RWRUH",
    "B06": "DARUH", "B07": "LBRUH", "B08": "SWRUH", "B09": "AZRUH", "B10": "SHRUH",
    "B11": "NRRUH", "B12": "TWRUH", "B13": "AQRUH", "B14": "RBRUH", "B15": "NDRUH",
    "B16": "BDRUH", "B17": "QRRUH", "B18": "TKRUH", "B19": "MURUH", "B21": "KRRUH",
    "B22": "OBJED", "B24": "SFJED", "B25": "RWAHS", "B26": "HAJED", "B27": "SARUH",
    "B28": "MAJED", "B30": "QARUH", "B31": "ANRUH", "B32": "FYJED", "B33": "HIRJED",
    "B34": "URRUH", "B35": "IRRUH", "B36": "PSJED", "B37": "SHMAK", "B38": "UHDMM",
    "B39": "HSRUH", "B23": "SLAHS",
}

# Ordered list of branch codes for sorting (B01, B02, … B39)
BRANCH_CODE_ORDER = [
    "B01","B02","B03","B04","B05","B06","B07","B08","B09","B10",
    "B11","B12","B13","B14","B15","B16","B17","B18","B19","B20",
    "B21","B22","B23","B24","B25","B26","B27","B28","B29","B30",
    "B31","B32","B33","B34","B35","B36","B37","B38","B39",
]

def branch_code_sort_key(name):
    """Sort key: by branch code number if name maps to a code, else alphabetical at end."""
    # name might be branch name (e.g. "NURUH") or branch code (e.g. "B01") or unknown
    # try reverse lookup: name → code
    for code, bname in BRANCH_CODE_NAME_MAP.items():
        if bname == name:
            try:
                return (BRANCH_CODE_ORDER.index(code), name)
            except ValueError:
                pass
    # maybe it IS a code already
    try:
        return (BRANCH_CODE_ORDER.index(name), name)
    except ValueError:
        return (9999, name)

# TERMINAL_BANK_MAP: terminal → Bank Name (always "Bank Al Bilad" for known terminals)
TERMINAL_BANK_MAP = {tid: "Bank Al Bilad" for tid in TERMINAL_BRANCH_CODE_MAP}
# Add legacy terminals not in branch map
_LEGACY = [
    "63189108","63189112","63189116","63189117","63189119","63189120",
    "63189167","63189168","63189169","63189491","63189492","63189497",
    "63189499","63189503","63189506","63933955","63933956","63933958",
    "63933959","63934016","63934018","63934019","63934020","63934021",
    "63934022","63934023","63934024","63934025",
]
for _t in _LEGACY:
    TERMINAL_BANK_MAP[_t] = "Bank Al Bilad"

# TERMINAL_BRANCH_LABEL_MAP: terminal → branch name label for column headers
TERMINAL_BRANCH_LABEL_MAP = {
    tid: BRANCH_CODE_NAME_MAP.get(code, f"#{tid}")
    for tid, code in TERMINAL_BRANCH_CODE_MAP.items()
}

# ==================== FILE READING & DETECTION ====================

def read_uploaded_file(uploaded_file):
    file_name = uploaded_file.name.lower()
    try:
        if file_name.endswith('.xlsx'):
            return pd.read_excel(uploaded_file, engine='openpyxl')
        elif file_name.endswith('.xls'):
            return pd.read_excel(uploaded_file, engine='xlrd')
        elif file_name.endswith('.csv'):
            try:
                uploaded_file.seek(0)
                df = pd.read_csv(uploaded_file, sep='\t')
                if df.shape[1] > 1:
                    return df
            except Exception:
                pass
            uploaded_file.seek(0)
            try:
                return pd.read_csv(uploaded_file)
            except Exception:
                uploaded_file.seek(0)
                return pd.read_csv(uploaded_file, sep=';')
        else:
            try:
                return pd.read_excel(uploaded_file, engine='openpyxl')
            except Exception:
                uploaded_file.seek(0)
                try:
                    return pd.read_excel(uploaded_file, engine='xlrd')
                except Exception:
                    uploaded_file.seek(0)
                    return pd.read_csv(uploaded_file)
    except Exception as e:
        raise Exception(f"Could not read file '{uploaded_file.name}'. Error: {str(e)}")


def detect_file_type(df):
    columns = [str(col).lower().strip() for col in df.columns]
    if any("payment method" in col for col in columns) and any("branch" in col for col in columns):
        return "foodics"
    if "terminal" in columns and "card name" in columns:
        return "geidea"
    try:
        all_values = df.astype(str).values.flatten()
        if any("Payment Method" in v for v in all_values) and any("Branch" in v for v in all_values):
            return "foodics"
    except Exception:
        pass
    return "unknown"


def parse_foodics_date_range(date_range_str):
    try:
        dates = re.findall(r"\d{4}-\d{2}-\d{2}", str(date_range_str))
        if len(dates) >= 2:
            start_date = pd.to_datetime(dates[0]).date()
            end_date   = pd.to_datetime(dates[1]).date()
            return pd.date_range(start=start_date, end=end_date, freq="D").date.tolist()
    except Exception:
        pass
    return []

# ==================== DATA PROCESSING ====================

def process_geidea_data(df):
    df = df.copy()
    df["Terminal"]  = df["Terminal"].astype(str).str.strip().str.replace(".0", "", regex=False)
    df["Bank Name"] = df["Terminal"].map(TERMINAL_BANK_MAP).fillna("Unknown Bank")
    df["Total"]     = df["Ter. Total Debit"].fillna(0) + df["Ter. Total Credit"].fillna(0)
    df["Total Debit"]        = df["Ter. Total Debit"]
    df["Total Credit"]       = df["Ter. Total Credit"]
    df["Total Debit Credit"] = df["Ter.Total Debit Credit"]
    date_col = next((col for col in df.columns if "date" in col.lower() and "recon" in col.lower()), None)
    df["Reconciliation Date"] = pd.to_datetime(df[date_col]).dt.date if date_col else None
    return df


def process_foodics_data(df):
    columns_lower = [str(col).lower().strip() for col in df.columns]
    if "payment method" in columns_lower and "branch" in columns_lower:
        df_clean   = df.copy()
        df_clean.columns = [str(col).strip() for col in df_clean.columns]
        date_range = ""
        dates      = []
    else:
        data_start = 0
        for idx, row in df.iterrows():
            if "Payment Method" in str(row.values):
                data_start = idx
                break
        df_clean = df.iloc[data_start:].reset_index(drop=True)
        df_clean.columns = df_clean.iloc[0]
        df_clean = df_clean[1:].reset_index(drop=True)
        df_clean.columns = [str(col).strip() for col in df_clean.columns]
        date_range = ""
        for idx, row in df.iterrows():
            if "Date Range" in str(row.values):
                date_range = row.iloc[1] if len(row) > 1 else ""
                break
        dates = parse_foodics_date_range(date_range)

    df_clean["Net Amount"]    = pd.to_numeric(df_clean["Net Amount"],    errors="coerce").fillna(0)
    df_clean["Amount"]        = pd.to_numeric(df_clean["Amount"],        errors="coerce").fillna(0)
    df_clean["Return Amount"] = pd.to_numeric(df_clean["Return Amount"], errors="coerce").fillna(0)
    df_clean["Count"]         = pd.to_numeric(df_clean["Count"],         errors="coerce").fillna(0).astype(int)
    df_clean["Report Date Range"] = date_range
    df_clean["Dates"] = [dates] * len(df_clean)
    return df_clean, dates

# ==================== SHARED EXCEL HELPERS ====================

def _apply_simplified_pivot_sheet(ws, df_subset, label, header_hex, sub_hex,
                                   tab_label="Detailed_Total_Only"):
    """
    Write a simplified pivot (Total only, no Avg) to ws.
    Rows = Bank Name × Card Scheme
    Cols = one column per BRANCH (terminals grouped by branch label, summed)
    Plus a GRAND TOTAL column at the end.
    Bottom: TOTAL row.
    Used for both current-day and previous-day tabs.
    """
    df_work = df_subset.copy()
    # Map each terminal to its branch label
    df_work["Branch Label"] = df_work["Terminal"].apply(
        lambda t: TERMINAL_BRANCH_LABEL_MAP.get(str(t).strip(), f"#{t}")
    )

    # Group by Branch Label (not individual terminal) + Bank Name + Card Name
    summary = df_work.groupby(["Branch Label", "Bank Name", "Card Name"]).agg(
        {"Total Debit Credit": "sum"}
    ).reset_index()

    branches     = sorted(summary["Branch Label"].unique(), key=branch_code_sort_key)
    banks        = sorted(summary["Bank Name"].unique(), key=lambda x: (x == "Unknown Bank", branch_code_sort_key(x)))
    card_schemes = sorted(summary["Card Name"].unique())

    pivot = {}
    for _, row in summary.iterrows():
        pivot[(row["Bank Name"], row["Card Name"], row["Branch Label"])] = row["Total Debit Credit"]

    col_totals = {b: sum(pivot.get((bk, c, b), 0) for bk in banks for c in card_schemes)
                  for b in branches}

    header_fill  = PatternFill(start_color=header_hex, end_color=header_hex, fill_type="solid")
    sub_fill     = PatternFill(start_color=sub_hex,    end_color=sub_hex,    fill_type="solid")
    unknown_fill = PatternFill(start_color="FF6B6B",   end_color="FF6B6B",   fill_type="solid")
    total_fill   = PatternFill(start_color="FFC000",   end_color="FFC000",   fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")
    right  = Alignment(horizontal="right")

    # ── Label header (label eg. "Today" or "Previous Day") ────────────────
    if label:
        label_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        total_header_cols = 2 + len(branches) + 1   # Bank + Card + branches + Grand
        ws.cell(row=1, column=1, value=label)
        ws.cell(row=1, column=1).fill  = label_fill
        ws.cell(row=1, column=1).font  = Font(color="FFFFFF", bold=True, size=12)
        ws.cell(row=1, column=1).alignment = center
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_header_cols)
        hdr_start = 2
    else:
        hdr_start = 1

    # ── Row hdr_start: column headers ────────────────────────────────────
    # Bank Name and Card Scheme span both header rows
    for c, v in [(1, "Bank Name"), (2, "Card Scheme")]:
        ws.cell(row=hdr_start,     column=c, value=v)
        ws.cell(row=hdr_start,     column=c).fill      = header_fill
        ws.cell(row=hdr_start,     column=c).font      = Font(color="FFFFFF", bold=True, size=10)
        ws.cell(row=hdr_start,     column=c).alignment = center
        ws.merge_cells(start_row=hdr_start, start_column=c,
                       end_row=hdr_start + 1, end_column=c)

    col_idx = 3
    for branch in branches:
        ws.cell(row=hdr_start, column=col_idx, value=branch)
        ws.cell(row=hdr_start, column=col_idx).fill      = header_fill
        ws.cell(row=hdr_start, column=col_idx).font      = Font(color="FFFFFF", bold=True, size=9)
        ws.cell(row=hdr_start, column=col_idx).alignment = center
        # sub-header row
        c2 = ws.cell(row=hdr_start + 1, column=col_idx, value="Total")
        c2.fill = sub_fill; c2.font = Font(bold=True, size=9); c2.alignment = center
        col_idx += 1

    # Grand Total column header
    grand_col = col_idx
    ws.cell(row=hdr_start, column=grand_col, value="GRAND TOTAL")
    ws.cell(row=hdr_start, column=grand_col).fill      = total_fill
    ws.cell(row=hdr_start, column=grand_col).font      = Font(bold=True, size=10)
    ws.cell(row=hdr_start, column=grand_col).alignment = center
    c2 = ws.cell(row=hdr_start + 1, column=grand_col, value="Total")
    c2.fill = total_fill; c2.font = Font(bold=True, size=9); c2.alignment = center

    # ── Data rows ─────────────────────────────────────────────────────────
    data_start_row = hdr_start + 2
    row_idx = data_start_row
    for bank in banks:
        for card in card_schemes:
            if not any((bank, card, br) in pivot for br in branches):
                continue
            is_unknown = bank == "Unknown Bank"
            for c, val in [(1, bank), (2, card)]:
                cell = ws.cell(row=row_idx, column=c, value=val)
                if is_unknown:
                    cell.fill = unknown_fill
                    cell.font = Font(bold=True, color="FFFFFF")

            grand_row = 0.0
            col = 3
            for branch in branches:
                val = pivot.get((bank, card, branch), 0)
                cell_t = ws.cell(row=row_idx, column=col, value=val)
                cell_t.number_format = "#,##0.00"
                cell_t.alignment     = right
                if is_unknown:
                    cell_t.fill = unknown_fill
                grand_row += val
                col += 1

            cell_gt = ws.cell(row=row_idx, column=grand_col, value=grand_row)
            cell_gt.number_format = "#,##0.00"
            cell_gt.font          = Font(bold=True)
            cell_gt.alignment     = right
            if is_unknown:
                cell_gt.fill = unknown_fill
            row_idx += 1

    # ── TOTAL row ─────────────────────────────────────────────────────────
    row_idx += 1
    ws.cell(row=row_idx, column=1, value="TOTAL").fill = total_fill
    ws.cell(row=row_idx, column=1).font = Font(bold=True, size=11)
    ws.cell(row=row_idx, column=2, value="").fill = total_fill

    grand_total = 0.0
    col = 3
    for branch in branches:
        val = col_totals[branch]
        cell = ws.cell(row=row_idx, column=col, value=val)
        cell.fill = total_fill; cell.font = Font(bold=True)
        cell.number_format = "#,##0.00"; cell.alignment = right
        grand_total += val
        col += 1

    cell_gt = ws.cell(row=row_idx, column=grand_col, value=grand_total)
    cell_gt.fill = total_fill; cell_gt.font = Font(bold=True, size=11)
    cell_gt.number_format = "#,##0.00"; cell_gt.alignment = right

    # ── Column widths ─────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 16
    for i in range(3, grand_col + 1):
        ws.column_dimensions[get_column_letter(i)].width = 14

    return len(branches)


# ==================== GEIDEA EXCEL FUNCTIONS ====================

def create_geidea_summary_file(df):
    summary = df.groupby(["Bank Name", "Card Name"]).agg({"Total": "sum"}).reset_index()
    summary["Sort"]      = summary["Bank Name"].apply(lambda x: 1 if x == "Unknown Bank" else 0)
    summary["BranchOrd"] = summary["Bank Name"].apply(branch_code_sort_key)
    summary = summary.sort_values(["Sort", "BranchOrd", "Card Name"]).drop(["Sort","BranchOrd"], axis=1)

    wb = Workbook(); ws = wb.active; ws.title = "Summary"
    header_fill  = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    unknown_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    total_fill   = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

    for col, header in enumerate(["Bank Name", "Card Scheme", "Total"], 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill; cell.font = Font(color="FFFFFF", bold=True, size=12)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    row_idx = 2
    for _, data in summary.iterrows():
        ws.cell(row=row_idx, column=1, value=data["Bank Name"])
        ws.cell(row=row_idx, column=2, value=data["Card Name"])
        c = ws.cell(row=row_idx, column=3, value=data["Total"])
        c.number_format = "#,##0.00"; c.alignment = Alignment(horizontal="right")
        if data["Bank Name"] == "Unknown Bank":
            for col in range(1, 4):
                ws.cell(row=row_idx, column=col).fill = unknown_fill
                ws.cell(row=row_idx, column=col).font = Font(bold=True, color="FFFFFF")
        row_idx += 1

    grand_total = summary["Total"].sum()
    row_idx += 1
    for col, val in {1: "GRAND TOTAL", 2: "ALL", 3: grand_total}.items():
        cell = ws.cell(row=row_idx, column=col, value=val)
        cell.fill = total_fill; cell.font = Font(bold=True, size=12)
    ws.cell(row=row_idx, column=3).number_format = "#,##0.00"

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 15
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf, summary, grand_total


def create_geidea_summary_by_date_file(df):
    if df["Reconciliation Date"].isna().all():
        return None, None, 0
    summary = df.groupby(["Reconciliation Date", "Bank Name", "Card Name"]).agg({"Total": "sum"}).reset_index()
    summary["Sort"]      = summary["Bank Name"].apply(lambda x: 1 if x == "Unknown Bank" else 0)
    summary["BranchOrd"] = summary["Bank Name"].apply(branch_code_sort_key)
    summary = summary.sort_values(["Reconciliation Date", "Sort", "BranchOrd", "Card Name"]).drop(["Sort","BranchOrd"], axis=1)

    wb = Workbook(); ws = wb.active; ws.title = "Summary_by_Date"
    header_fill   = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    date_fill     = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    unknown_fill  = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    total_fill    = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    subtotal_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

    for col, header in enumerate(["Date", "Bank Name", "Card Scheme", "Total"], 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill; cell.font = Font(color="FFFFFF", bold=True, size=12)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    row_idx = 2; current_date = None; date_totals = {}
    for _, data in summary.iterrows():
        date_val = data["Reconciliation Date"]
        date_str = date_val.strftime("%A/%d/%b/%Y") if hasattr(date_val, "strftime") else str(date_val)
        if date_val != current_date:
            if current_date is not None:
                row_idx += 1
                ws.cell(row=row_idx, column=2, value="DATE SUBTOTAL")
                c = ws.cell(row=row_idx, column=4, value=date_totals[current_date])
                c.number_format = "#,##0.00"; c.font = Font(bold=True)
                for col in range(1, 5): ws.cell(row=row_idx, column=col).fill = subtotal_fill
                row_idx += 1
            ws.cell(row=row_idx, column=1, value=date_str)
            ws.cell(row=row_idx, column=1).fill = date_fill
            ws.cell(row=row_idx, column=1).font = Font(color="FFFFFF", bold=True, size=11)
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=4)
            row_idx += 1; current_date = date_val; date_totals[current_date] = 0

        ws.cell(row=row_idx, column=2, value=data["Bank Name"])
        ws.cell(row=row_idx, column=3, value=data["Card Name"])
        c = ws.cell(row=row_idx, column=4, value=data["Total"])
        c.number_format = "#,##0.00"; c.alignment = Alignment(horizontal="right")
        if data["Bank Name"] == "Unknown Bank":
            for col in range(2, 5):
                ws.cell(row=row_idx, column=col).fill = unknown_fill
                ws.cell(row=row_idx, column=col).font = Font(bold=True, color="FFFFFF")
        date_totals[current_date] += data["Total"]; row_idx += 1

    if current_date is not None:
        row_idx += 1
        ws.cell(row=row_idx, column=2, value="DATE SUBTOTAL")
        c = ws.cell(row=row_idx, column=4, value=date_totals[current_date])
        c.number_format = "#,##0.00"; c.font = Font(bold=True)
        for col in range(1, 5): ws.cell(row=row_idx, column=col).fill = subtotal_fill
        row_idx += 1

    row_idx += 1; grand_total = summary["Total"].sum()
    for col, val in {2: "GRAND TOTAL", 3: "ALL DATES", 4: grand_total}.items():
        cell = ws.cell(row=row_idx, column=col, value=val)
        cell.fill = total_fill; cell.font = Font(bold=True, size=12)
    ws.cell(row=row_idx, column=4).number_format = "#,##0.00"
    for ltr, w in zip("ABCD", [20, 20, 18, 15]): ws.column_dimensions[ltr].width = w
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf, summary, len(summary["Reconciliation Date"].unique())


def create_geidea_detailed_file(df):
    summary = df.groupby(["Terminal", "Bank Name", "Card Name"]).agg({
        "Total Debit": "sum", "Total Credit": "sum", "Total Debit Credit": "sum"
    }).reset_index()
    terminals    = sorted(summary["Terminal"].unique(), key=lambda t: branch_code_sort_key(TERMINAL_BRANCH_LABEL_MAP.get(t, t)))
    banks        = sorted(summary["Bank Name"].unique(), key=lambda x: (x == "Unknown Bank", branch_code_sort_key(x)))
    card_schemes = sorted(summary["Card Name"].unique())

    rows = []
    for bank in banks:
        for card in card_schemes:
            bc = summary[(summary["Bank Name"] == bank) & (summary["Card Name"] == card)]
            if bc.empty: continue
            row = {"Bank Name": bank, "Card Scheme": card}
            for term in terminals:
                td = bc[bc["Terminal"] == term]
                row[f"{term}_Debit"]  = td["Total Debit"].values[0]        if not td.empty else 0
                row[f"{term}_Credit"] = td["Total Credit"].values[0]       if not td.empty else 0
                row[f"{term}_Total"]  = td["Total Debit Credit"].values[0] if not td.empty else 0
            rows.append(row)
    for label in ["TOTAL", "AVG"]:
        row = {"Bank Name": label, "Card Scheme": "ALL"}
        for term in terminals:
            td = summary[summary["Terminal"] == term]
            row[f"{term}_Debit"]  = round(td["Total Debit"].sum()        if label=="TOTAL" else td["Total Debit"].mean(),        2)
            row[f"{term}_Credit"] = round(td["Total Credit"].sum()       if label=="TOTAL" else td["Total Credit"].mean(),       2)
            row[f"{term}_Total"]  = round(td["Total Debit Credit"].sum() if label=="TOTAL" else td["Total Debit Credit"].mean(), 2)
        rows.append(row)

    wb = Workbook(); ws = wb.active; ws.title = "Detailed"
    header_fill  = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    sub_fill     = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
    unknown_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")

    for c, v in [(1, "Bank Name"), (2, "Card Scheme")]:
        cell = ws.cell(row=1, column=c, value=v)
        cell.fill = header_fill; cell.font = Font(color="FFFFFF", bold=True, size=9); cell.alignment = center

    col_idx = 3
    for term in terminals:
        ws.cell(row=1, column=col_idx, value=TERMINAL_BRANCH_LABEL_MAP.get(term, f"#{term}")).fill = header_fill
        ws.cell(row=1, column=col_idx).font = Font(color="FFFFFF", bold=True, size=9)
        ws.cell(row=1, column=col_idx).alignment = center
        ws.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx + 2)
        for lbl, off in [("Debit", 0), ("Credit", 1), ("Total", 2)]:
            c2 = ws.cell(row=2, column=col_idx + off, value=lbl)
            c2.fill = sub_fill; c2.font = Font(bold=True, size=8); c2.alignment = center
        col_idx += 3

    for r_idx, row_data in enumerate(rows, 3):
        bv, cv = row_data["Bank Name"], row_data["Card Scheme"]
        for c, val in [(1, bv), (2, cv)]:
            cell = ws.cell(row=r_idx, column=c, value=val)
            if bv == "Unknown Bank":
                cell.fill = unknown_fill; cell.font = Font(bold=True, color="FFFFFF")
            elif bv in ["TOTAL", "AVG"]:
                cell.fill = PatternFill(start_color="E0E0E0", fill_type="solid"); cell.font = Font(bold=True)
        col_idx = 3
        for term in terminals:
            for off, key in enumerate(["Debit", "Credit", "Total"]):
                cell = ws.cell(row=r_idx, column=col_idx + off, value=row_data[f"{term}_{key}"])
                cell.number_format = "#,##0.00"
            col_idx += 3

    ws.column_dimensions["A"].width = 18; ws.column_dimensions["B"].width = 15
    for i in range(3, col_idx): ws.column_dimensions[get_column_letter(i)].width = 11
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf, len(terminals)


def _extract_date_label(df, subtract_day=False):
    """
    Extract the reconciliation date from a processed df as a short string like '24-Feb'.
    If subtract_day=True, subtracts 1 day (for previous-day rows that share the same
    Reconciliation Date as today because the timestamp rolled over midnight).
    """
    import datetime
    try:
        dates = df["Reconciliation Date"].dropna().unique()
        if len(dates) > 0:
            d = sorted(dates)[-1]   # take the latest date
            if hasattr(d, "strftime"):
                if subtract_day:
                    d = d - datetime.timedelta(days=1)
                return d.strftime("%d-%b")   # e.g. "24-Feb"
        return None
    except Exception:
        return None


def create_geidea_detailed_totals_only(df_today, df_prev=None):
    """
    Simplified Geidea: Total only (no Avg) per terminal.
    Sheet tabs named by actual date (e.g. '24-Feb').
    Previous-day rows share the same Reconciliation Date as today (the timestamp
    rolled over midnight), so we subtract 1 day for their label.
    """
    import datetime

    # Today label straight from data
    today_label = _extract_date_label(df_today, subtract_day=False)
    if today_label is None:
        today_label = datetime.date.today().strftime("%d-%b")

    prev_label = None
    if df_prev is not None and len(df_prev) > 0:
        # Previous-day rows have the SAME Reconciliation Date as today
        # (Geidea stamps the reconciliation date, not the transaction date)
        # so we subtract 1 day to get the correct label
        prev_label = _extract_date_label(df_prev, subtract_day=True)
        if prev_label is None:
            try:
                d = datetime.date.today() - datetime.timedelta(days=1)
                prev_label = d.strftime("%d-%b")
            except Exception:
                prev_label = "Prev Day"

    wb = Workbook()

    # ── Today sheet (named by actual date) ───────────────────────────────
    ws_today = wb.active
    ws_today.title = today_label
    _apply_simplified_pivot_sheet(
        ws_today, df_today,
        label=today_label,
        header_hex="366092", sub_hex="B8CCE4"
    )

    # ── Previous Day sheet (named by its actual date) ─────────────────────
    if prev_label is not None:
        ws_prev = wb.create_sheet(title=prev_label)
        _apply_simplified_pivot_sheet(
            ws_prev, df_prev,
            label=prev_label,
            header_hex="1F4E78", sub_hex="9DC3E6"
        )

    n_terminals = len(df_today["Terminal"].unique()) if "Terminal" in df_today.columns else 0
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf, n_terminals


def create_geidea_detailed_by_date_file(df):
    if df["Reconciliation Date"].isna().all():
        return None, 0, 0
    summary = df.groupby(["Reconciliation Date", "Terminal", "Bank Name", "Card Name"]).agg({
        "Total Debit": "sum", "Total Credit": "sum", "Total Debit Credit": "sum"
    }).reset_index()
    dates        = sorted(summary["Reconciliation Date"].unique())
    terminals    = sorted(summary["Terminal"].unique(), key=lambda t: branch_code_sort_key(TERMINAL_BRANCH_LABEL_MAP.get(t, t)))
    banks        = sorted(summary["Bank Name"].unique(), key=lambda x: (x == "Unknown Bank", branch_code_sort_key(x)))
    card_schemes = sorted(summary["Card Name"].unique())

    rows = []
    for bank in banks:
        for card in card_schemes:
            bc = summary[(summary["Bank Name"] == bank) & (summary["Card Name"] == card)]
            if bc.empty: continue
            row = {"Bank Name": bank, "Card Scheme": card}
            for date in dates:
                dd = bc[bc["Reconciliation Date"] == date]
                for term in terminals:
                    td = dd[dd["Terminal"] == term]
                    row[f"{date}_{term}_Debit"]  = td["Total Debit"].values[0]        if not td.empty else 0
                    row[f"{date}_{term}_Credit"] = td["Total Credit"].values[0]       if not td.empty else 0
                    row[f"{date}_{term}_Total"]  = td["Total Debit Credit"].values[0] if not td.empty else 0
            rows.append(row)
    for label in ["TOTAL", "AVG"]:
        row = {"Bank Name": label, "Card Scheme": "ALL"}
        for date in dates:
            dd = summary[summary["Reconciliation Date"] == date]
            for term in terminals:
                td = dd[dd["Terminal"] == term]
                row[f"{date}_{term}_Debit"]  = round(td["Total Debit"].sum()        if label=="TOTAL" else td["Total Debit"].mean(),        2)
                row[f"{date}_{term}_Credit"] = round(td["Total Credit"].sum()       if label=="TOTAL" else td["Total Credit"].mean(),       2)
                row[f"{date}_{term}_Total"]  = round(td["Total Debit Credit"].sum() if label=="TOTAL" else td["Total Debit Credit"].mean(), 2)
        rows.append(row)

    wb = Workbook(); ws = wb.active; ws.title = "Detailed_by_Date"
    date_fill    = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_fill  = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    sub_fill     = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
    unknown_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")

    for c, v in [(1, "Bank Name"), (2, "Card Scheme")]:
        cell = ws.cell(row=1, column=c, value=v)
        cell.fill = header_fill; cell.font = Font(color="FFFFFF", bold=True, size=9); cell.alignment = center

    col_idx = 3
    for date in dates:
        end_col = col_idx + (len(terminals) * 3) - 1
        ws.cell(row=1, column=col_idx, value=date.strftime("%A/%d/%b/%Y"))
        ws.cell(row=1, column=col_idx).fill = date_fill
        ws.cell(row=1, column=col_idx).font = Font(color="FFFFFF", bold=True, size=11)
        ws.cell(row=1, column=col_idx).alignment = center
        ws.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=end_col)
        term_col = col_idx
        for term in terminals:
            ws.cell(row=2, column=term_col, value=TERMINAL_BRANCH_LABEL_MAP.get(term, f"#{term}")).fill = header_fill
            ws.cell(row=2, column=term_col).font = Font(color="FFFFFF", bold=True, size=9)
            ws.cell(row=2, column=term_col).alignment = center
            ws.merge_cells(start_row=2, start_column=term_col, end_row=2, end_column=term_col + 2)
            for lbl, off in [("Debit", 0), ("Credit", 1), ("Total", 2)]:
                c2 = ws.cell(row=3, column=term_col + off, value=lbl)
                c2.fill = sub_fill; c2.font = Font(bold=True, size=8); c2.alignment = center
            term_col += 3
        col_idx = end_col + 1

    for r_idx, row_data in enumerate(rows, 4):
        bv, cv = row_data["Bank Name"], row_data["Card Scheme"]
        for c, val in [(1, bv), (2, cv)]:
            cell = ws.cell(row=r_idx, column=c, value=val)
            if bv == "Unknown Bank":
                cell.fill = unknown_fill; cell.font = Font(bold=True, color="FFFFFF")
            elif bv in ["TOTAL", "AVG"]:
                cell.fill = PatternFill(start_color="E0E0E0", fill_type="solid"); cell.font = Font(bold=True)
        col_idx = 3
        for date in dates:
            for term in terminals:
                for off, key in enumerate(["Debit", "Credit", "Total"]):
                    cell = ws.cell(row=r_idx, column=col_idx + off, value=row_data[f"{date}_{term}_{key}"])
                    cell.number_format = "#,##0.00"
                col_idx += 3

    ws.column_dimensions["A"].width = 18; ws.column_dimensions["B"].width = 15
    for i in range(3, col_idx): ws.column_dimensions[get_column_letter(i)].width = 11
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf, len(dates), len(terminals)


# ==================== FOODICS FUNCTIONS ====================

def create_foodics_summary_by_branch(df):
    summary = df.groupby(["Branch", "Payment Method"]).agg({
        "Net Amount": "sum", "Amount": "sum", "Return Amount": "sum", "Count": "sum"
    }).reset_index()
    summary["BranchOrd"] = summary["Branch"].apply(branch_code_sort_key)
    summary = summary.sort_values(["BranchOrd", "Net Amount"], ascending=[True, False]).drop("BranchOrd", axis=1)

    wb = Workbook(); ws = wb.active; ws.title = "Summary_by_Branch"
    header_fill   = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    branch_fill   = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    subtotal_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    total_fill    = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")

    for col, header in enumerate(["Branch", "Payment Method", "Net Amount", "Amount", "Returns", "Count"], 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill; cell.font = Font(color="FFFFFF", bold=True, size=11); cell.alignment = center

    row_idx = 2; current_branch = None; branch_totals = {}
    for _, data in summary.iterrows():
        branch = data["Branch"]
        if branch != current_branch:
            if current_branch is not None:
                row_idx += 1; bt = branch_totals[current_branch]
                for col, val in [(2,"BRANCH SUBTOTAL"),(3,bt["net"]),(4,bt["amount"]),(5,bt["returns"]),(6,bt["count"])]:
                    ws.cell(row=row_idx, column=col, value=val)
                for col in range(1, 7):
                    ws.cell(row=row_idx, column=col).fill = subtotal_fill
                    ws.cell(row=row_idx, column=col).font = Font(bold=True)
                    if col >= 3: ws.cell(row=row_idx, column=col).number_format = "#,##0.00"
                row_idx += 1
            ws.cell(row=row_idx, column=1, value=branch)
            ws.cell(row=row_idx, column=1).fill = branch_fill
            ws.cell(row=row_idx, column=1).font = Font(color="FFFFFF", bold=True, size=10)
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=6)
            row_idx += 1; current_branch = branch
            branch_totals[current_branch] = {"net": 0, "amount": 0, "returns": 0, "count": 0}

        ws.cell(row=row_idx, column=2, value=data["Payment Method"])
        ws.cell(row=row_idx, column=3, value=data["Net Amount"])
        ws.cell(row=row_idx, column=4, value=data["Amount"])
        ws.cell(row=row_idx, column=5, value=data["Return Amount"])
        ws.cell(row=row_idx, column=6, value=data["Count"])
        for col in range(3, 7): ws.cell(row=row_idx, column=col).number_format = "#,##0.00" if col < 6 else "#,##0"
        branch_totals[current_branch]["net"]     += data["Net Amount"]
        branch_totals[current_branch]["amount"]  += data["Amount"]
        branch_totals[current_branch]["returns"] += data["Return Amount"]
        branch_totals[current_branch]["count"]   += data["Count"]
        row_idx += 1

    if current_branch is not None:
        row_idx += 1; bt = branch_totals[current_branch]
        for col, val in [(2,"BRANCH SUBTOTAL"),(3,bt["net"]),(4,bt["amount"]),(5,bt["returns"]),(6,bt["count"])]:
            ws.cell(row=row_idx, column=col, value=val)
        for col in range(1, 7):
            ws.cell(row=row_idx, column=col).fill = subtotal_fill
            ws.cell(row=row_idx, column=col).font = Font(bold=True)
            if col >= 3: ws.cell(row=row_idx, column=col).number_format = "#,##0.00"
        row_idx += 1

    row_idx += 1
    for col, val in [(2,"GRAND TOTAL"),(3,summary["Net Amount"].sum()),(4,summary["Amount"].sum()),
                     (5,summary["Return Amount"].sum()),(6,summary["Count"].sum())]:
        cell = ws.cell(row=row_idx, column=col, value=val)
        cell.fill = total_fill; cell.font = Font(bold=True, size=12)
        if col >= 3: cell.number_format = "#,##0.00"

    for ltr, w in zip("ABCDEF", [15, 25, 15, 15, 15, 12]): ws.column_dimensions[ltr].width = w
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf, summary, len(summary["Branch"].unique())


def create_foodics_summary_by_payment_method(df):
    summary = df.groupby(["Payment Method"]).agg({
        "Net Amount": "sum", "Amount": "sum", "Return Amount": "sum", "Count": "sum"
    }).reset_index().sort_values("Net Amount", ascending=False)

    wb = Workbook(); ws = wb.active; ws.title = "Summary_by_Payment"
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    total_fill  = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")

    for col, header in enumerate(["Payment Method", "Net Amount", "Amount", "Returns", "Count"], 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill; cell.font = Font(color="FFFFFF", bold=True, size=11); cell.alignment = center

    row_idx = 2
    for _, data in summary.iterrows():
        ws.cell(row=row_idx, column=1, value=data["Payment Method"])
        ws.cell(row=row_idx, column=2, value=data["Net Amount"])
        ws.cell(row=row_idx, column=3, value=data["Amount"])
        ws.cell(row=row_idx, column=4, value=data["Return Amount"])
        ws.cell(row=row_idx, column=5, value=data["Count"])
        for col in range(2, 6): ws.cell(row=row_idx, column=col).number_format = "#,##0.00" if col < 5 else "#,##0"
        row_idx += 1

    row_idx += 1
    for col, val in [(1,"GRAND TOTAL"),(2,summary["Net Amount"].sum()),(3,summary["Amount"].sum()),
                     (4,summary["Return Amount"].sum()),(5,summary["Count"].sum())]:
        cell = ws.cell(row=row_idx, column=col, value=val)
        cell.fill = total_fill; cell.font = Font(bold=True, size=12)
        if col >= 2: cell.number_format = "#,##0.00"

    for ltr, w in zip("ABCDE", [30, 15, 15, 15, 12]): ws.column_dimensions[ltr].width = w
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf, summary


def _foodics_net_only_pivot(df, row_field, col_field, row_label, col_label,
                             header_hex, sub_hex):
    """Simplified Foodics pivot — Net Amount only, Total column per branch/payment method."""
    summary  = df.groupby([row_field, col_field]).agg({"Net Amount": "sum"}).reset_index()
    def _branch_sort(vals, field_name):
        if field_name == "Branch":
            return sorted(vals, key=branch_code_sort_key)
        return sorted(vals)

    row_keys = _branch_sort(summary[row_field].unique(), row_field)
    col_keys = _branch_sort(summary[col_field].unique(), col_field)
    pivot    = {(r[row_field], r[col_field]): r["Net Amount"] for _, r in summary.iterrows()}
    col_totals = {ck: sum(pivot.get((rk, ck), 0) for rk in row_keys) for ck in col_keys}
    n_rows = len(row_keys)

    wb = Workbook(); ws = wb.active
    ws.title = f"Det_{col_field[:10]}_NetOnly"

    header_fill = PatternFill(start_color=header_hex, end_color=header_hex, fill_type="solid")
    sub_fill    = PatternFill(start_color=sub_hex,    end_color=sub_hex,    fill_type="solid")
    total_fill  = PatternFill(start_color="FFC000",   end_color="FFC000",   fill_type="solid")
    avg_fill    = PatternFill(start_color="E0E0E0",   end_color="E0E0E0",   fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")
    right  = Alignment(horizontal="right")

    # Row 1 & 2 headers — row_label spans both rows, each col_key = 1 col (Total only)
    ws.cell(row=1, column=1, value=row_label)
    ws.cell(row=1, column=1).fill = header_fill
    ws.cell(row=1, column=1).font = Font(color="FFFFFF", bold=True, size=10)
    ws.cell(row=1, column=1).alignment = center
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)

    col_idx = 2
    for ck in col_keys:
        ws.cell(row=1, column=col_idx, value=ck)
        ws.cell(row=1, column=col_idx).fill      = header_fill
        ws.cell(row=1, column=col_idx).font      = Font(color="FFFFFF", bold=True, size=9)
        ws.cell(row=1, column=col_idx).alignment = center
        c2 = ws.cell(row=2, column=col_idx, value="Total")
        c2.fill = sub_fill; c2.font = Font(bold=True, size=9); c2.alignment = center
        col_idx += 1

    grand_col = col_idx
    ws.cell(row=1, column=grand_col, value="GRAND TOTAL")
    ws.cell(row=1, column=grand_col).fill      = total_fill
    ws.cell(row=1, column=grand_col).font      = Font(bold=True, size=10)
    ws.cell(row=1, column=grand_col).alignment = center
    c2 = ws.cell(row=2, column=grand_col, value="Total")
    c2.fill = total_fill; c2.font = Font(bold=True, size=9); c2.alignment = center

    # Data rows
    row_idx = 3
    for rk in row_keys:
        ws.cell(row=row_idx, column=1, value=rk)
        grand_row = 0.0; col = 2
        for ck in col_keys:
            val = pivot.get((rk, ck), 0)
            ws.cell(row=row_idx, column=col, value=val).number_format = "#,##0.00"
            ws.cell(row=row_idx, column=col).alignment = right
            grand_row += val; col += 1
        ws.cell(row=row_idx, column=grand_col, value=grand_row).number_format = "#,##0.00"
        ws.cell(row=row_idx, column=grand_col).font      = Font(bold=True)
        ws.cell(row=row_idx, column=grand_col).alignment = right
        row_idx += 1

    # TOTAL row
    row_idx += 1
    ws.cell(row=row_idx, column=1, value="TOTAL").fill = total_fill
    ws.cell(row=row_idx, column=1).font = Font(bold=True, size=11)
    grand_total = 0.0; col = 2
    for ck in col_keys:
        val = col_totals[ck]
        cell = ws.cell(row=row_idx, column=col, value=val)
        cell.fill = total_fill; cell.font = Font(bold=True)
        cell.number_format = "#,##0.00"; cell.alignment = right
        grand_total += val; col += 1
    cell_gt = ws.cell(row=row_idx, column=grand_col, value=grand_total)
    cell_gt.fill = total_fill; cell_gt.font = Font(bold=True, size=11)
    cell_gt.number_format = "#,##0.00"; cell_gt.alignment = right

    # AVG row
    row_idx += 1
    ws.cell(row=row_idx, column=1, value="AVG").fill = avg_fill
    ws.cell(row=row_idx, column=1).font = Font(bold=True, size=11)
    col = 2
    for ck in col_keys:
        val = round(col_totals[ck] / n_rows, 2) if n_rows else 0
        cell = ws.cell(row=row_idx, column=col, value=val)
        cell.fill = avg_fill; cell.font = Font(bold=True)
        cell.number_format = "#,##0.00"; cell.alignment = right
        col += 1
    overall_avg = round(grand_total / n_rows, 2) if n_rows else 0
    cell_ga = ws.cell(row=row_idx, column=grand_col, value=overall_avg)
    cell_ga.fill = avg_fill; cell_ga.font = Font(bold=True, size=11)
    cell_ga.number_format = "#,##0.00"; cell_ga.alignment = right

    ws.column_dimensions["A"].width = 30
    for i in range(2, grand_col + 1): ws.column_dimensions[get_column_letter(i)].width = 14

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf, len(col_keys), len(row_keys)


def create_foodics_detailed_by_branch(df):
    summary = df.groupby(["Payment Method", "Branch"]).agg({
        "Net Amount": "sum", "Amount": "sum", "Return Amount": "sum", "Count": "sum"
    }).reset_index()
    payment_methods = sorted(summary["Payment Method"].unique())
    branches = sorted(summary["Branch"].unique(), key=branch_code_sort_key)

    wb = Workbook(); ws = wb.active; ws.title = "Detailed_by_Branch"
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    sub_fill    = PatternFill(start_color="A5D6A7", end_color="A5D6A7", fill_type="solid")
    total_fill  = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    avg_fill    = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")

    ws.cell(row=1, column=1, value="Payment Method")
    ws.cell(row=1, column=1).fill = header_fill
    ws.cell(row=1, column=1).font = Font(color="FFFFFF", bold=True, size=10)
    ws.cell(row=1, column=1).alignment = center

    col_idx = 2
    for branch in branches:
        ws.cell(row=1, column=col_idx, value=branch)
        ws.cell(row=1, column=col_idx).fill = header_fill
        ws.cell(row=1, column=col_idx).font = Font(color="FFFFFF", bold=True, size=9)
        ws.cell(row=1, column=col_idx).alignment = center
        ws.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx + 3)
        for lbl, off in [("Net Amount", 0), ("Amount", 1), ("Returns", 2), ("Count", 3)]:
            c2 = ws.cell(row=2, column=col_idx + off, value=lbl)
            c2.fill = sub_fill; c2.font = Font(bold=True, size=8); c2.alignment = center
        col_idx += 4

    ws.cell(row=1, column=col_idx, value="TOTAL").fill = total_fill
    ws.cell(row=1, column=col_idx).font = Font(bold=True, size=10); ws.cell(row=1, column=col_idx).alignment = center
    ws.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx + 3)
    for lbl, off in [("Net Amount", 0), ("Amount", 1), ("Returns", 2), ("Count", 3)]:
        c2 = ws.cell(row=2, column=col_idx + off, value=lbl)
        c2.fill = total_fill; c2.font = Font(bold=True, size=8); c2.alignment = center
    total_col_start = col_idx; col_idx += 4

    row_idx = 3
    for pm in payment_methods:
        ws.cell(row=row_idx, column=1, value=pm)
        pm_data = summary[summary["Payment Method"] == pm]
        c = 2; row_net = row_amt = row_ret = row_cnt = 0
        for branch in branches:
            bd = pm_data[pm_data["Branch"] == branch]
            net = bd["Net Amount"].values[0]    if not bd.empty else 0
            amt = bd["Amount"].values[0]         if not bd.empty else 0
            ret = bd["Return Amount"].values[0]  if not bd.empty else 0
            cnt = int(bd["Count"].values[0])     if not bd.empty else 0
            for off, val in enumerate([net, amt, ret, cnt]):
                cell = ws.cell(row=row_idx, column=c + off, value=val)
                cell.number_format = "#,##0.00" if off < 3 else "#,##0"
            row_net += net; row_amt += amt; row_ret += ret; row_cnt += cnt; c += 4
        for off, val in enumerate([row_net, row_amt, row_ret, row_cnt]):
            cell = ws.cell(row=row_idx, column=total_col_start + off, value=val)
            cell.number_format = "#,##0.00" if off < 3 else "#,##0"; cell.font = Font(bold=True)
        row_idx += 1

    row_idx += 1
    ws.cell(row=row_idx, column=1, value="TOTAL").fill = total_fill
    ws.cell(row=row_idx, column=1).font = Font(bold=True, size=11)
    c = 2
    for branch in branches:
        bd = summary[summary["Branch"] == branch]
        for off, cn in enumerate(["Net Amount", "Amount", "Return Amount", "Count"]):
            cell = ws.cell(row=row_idx, column=c + off, value=bd[cn].sum())
            cell.fill = total_fill; cell.font = Font(bold=True)
            cell.number_format = "#,##0.00" if off < 3 else "#,##0"
        c += 4
    for off, cn in enumerate(["Net Amount", "Amount", "Return Amount", "Count"]):
        cell = ws.cell(row=row_idx, column=total_col_start + off, value=summary[cn].sum())
        cell.fill = total_fill; cell.font = Font(bold=True, size=11)
        cell.number_format = "#,##0.00" if off < 3 else "#,##0"

    row_idx += 1
    ws.cell(row=row_idx, column=1, value="AVG").fill = avg_fill
    ws.cell(row=row_idx, column=1).font = Font(bold=True, size=11)
    c = 2
    for branch in branches:
        bd = summary[summary["Branch"] == branch]
        for off, cn in enumerate(["Net Amount", "Amount", "Return Amount", "Count"]):
            val = round(bd[cn].mean(), 2) if not bd.empty else 0
            cell = ws.cell(row=row_idx, column=c + off, value=val)
            cell.fill = avg_fill; cell.font = Font(bold=True)
            cell.number_format = "#,##0.00" if off < 3 else "#,##0"
        c += 4
    for off, cn in enumerate(["Net Amount", "Amount", "Return Amount", "Count"]):
        cell = ws.cell(row=row_idx, column=total_col_start + off, value=round(summary[cn].mean(), 2))
        cell.fill = avg_fill; cell.font = Font(bold=True); cell.number_format = "#,##0.00" if off < 3 else "#,##0"

    ws.column_dimensions["A"].width = 30
    for i in range(2, col_idx): ws.column_dimensions[get_column_letter(i)].width = 13
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf, len(branches), len(payment_methods)


def create_foodics_detailed_by_payment_method(df):
    summary = df.groupby(["Branch", "Payment Method"]).agg({
        "Net Amount": "sum", "Amount": "sum", "Return Amount": "sum", "Count": "sum"
    }).reset_index()
    branches        = sorted(summary["Branch"].unique(), key=branch_code_sort_key)
    payment_methods = sorted(summary["Payment Method"].unique())

    wb = Workbook(); ws = wb.active; ws.title = "Detailed_by_PayMethod"
    header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    sub_fill    = PatternFill(start_color="90CAF9", end_color="90CAF9", fill_type="solid")
    total_fill  = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    avg_fill    = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")

    ws.cell(row=1, column=1, value="Branch")
    ws.cell(row=1, column=1).fill = header_fill
    ws.cell(row=1, column=1).font = Font(color="FFFFFF", bold=True, size=10)
    ws.cell(row=1, column=1).alignment = center

    col_idx = 2
    for pm in payment_methods:
        ws.cell(row=1, column=col_idx, value=pm)
        ws.cell(row=1, column=col_idx).fill = header_fill
        ws.cell(row=1, column=col_idx).font = Font(color="FFFFFF", bold=True, size=9)
        ws.cell(row=1, column=col_idx).alignment = center
        ws.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx + 3)
        for lbl, off in [("Net Amount", 0), ("Amount", 1), ("Returns", 2), ("Count", 3)]:
            c2 = ws.cell(row=2, column=col_idx + off, value=lbl)
            c2.fill = sub_fill; c2.font = Font(bold=True, size=8); c2.alignment = center
        col_idx += 4

    ws.cell(row=1, column=col_idx, value="TOTAL").fill = total_fill
    ws.cell(row=1, column=col_idx).font = Font(bold=True, size=10); ws.cell(row=1, column=col_idx).alignment = center
    ws.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx + 3)
    for lbl, off in [("Net Amount", 0), ("Amount", 1), ("Returns", 2), ("Count", 3)]:
        c2 = ws.cell(row=2, column=col_idx + off, value=lbl)
        c2.fill = total_fill; c2.font = Font(bold=True, size=8); c2.alignment = center
    total_col_start = col_idx; col_idx += 4

    row_idx = 3
    for branch in branches:
        ws.cell(row=row_idx, column=1, value=branch)
        branch_data = summary[summary["Branch"] == branch]
        c = 2; row_net = row_amt = row_ret = row_cnt = 0
        for pm in payment_methods:
            pd_ = branch_data[branch_data["Payment Method"] == pm]
            net = pd_["Net Amount"].values[0]   if not pd_.empty else 0
            amt = pd_["Amount"].values[0]        if not pd_.empty else 0
            ret = pd_["Return Amount"].values[0] if not pd_.empty else 0
            cnt = int(pd_["Count"].values[0])    if not pd_.empty else 0
            for off, val in enumerate([net, amt, ret, cnt]):
                cell = ws.cell(row=row_idx, column=c + off, value=val)
                cell.number_format = "#,##0.00" if off < 3 else "#,##0"
            row_net += net; row_amt += amt; row_ret += ret; row_cnt += cnt; c += 4
        for off, val in enumerate([row_net, row_amt, row_ret, row_cnt]):
            cell = ws.cell(row=row_idx, column=total_col_start + off, value=val)
            cell.number_format = "#,##0.00" if off < 3 else "#,##0"; cell.font = Font(bold=True)
        row_idx += 1

    row_idx += 1
    ws.cell(row=row_idx, column=1, value="TOTAL").fill = total_fill
    ws.cell(row=row_idx, column=1).font = Font(bold=True, size=11)
    c = 2
    for pm in payment_methods:
        pd_ = summary[summary["Payment Method"] == pm]
        for off, cn in enumerate(["Net Amount", "Amount", "Return Amount", "Count"]):
            cell = ws.cell(row=row_idx, column=c + off, value=pd_[cn].sum())
            cell.fill = total_fill; cell.font = Font(bold=True)
            cell.number_format = "#,##0.00" if off < 3 else "#,##0"
        c += 4
    for off, cn in enumerate(["Net Amount", "Amount", "Return Amount", "Count"]):
        cell = ws.cell(row=row_idx, column=total_col_start + off, value=summary[cn].sum())
        cell.fill = total_fill; cell.font = Font(bold=True, size=11)
        cell.number_format = "#,##0.00" if off < 3 else "#,##0"

    row_idx += 1
    ws.cell(row=row_idx, column=1, value="AVG").fill = avg_fill
    ws.cell(row=row_idx, column=1).font = Font(bold=True, size=11)
    c = 2
    for pm in payment_methods:
        pd_ = summary[summary["Payment Method"] == pm]
        for off, cn in enumerate(["Net Amount", "Amount", "Return Amount", "Count"]):
            val = round(pd_[cn].mean(), 2) if not pd_.empty else 0
            cell = ws.cell(row=row_idx, column=c + off, value=val)
            cell.fill = avg_fill; cell.font = Font(bold=True)
            cell.number_format = "#,##0.00" if off < 3 else "#,##0"
        c += 4
    for off, cn in enumerate(["Net Amount", "Amount", "Return Amount", "Count"]):
        cell = ws.cell(row=row_idx, column=total_col_start + off, value=round(summary[cn].mean(), 2))
        cell.fill = avg_fill; cell.font = Font(bold=True); cell.number_format = "#,##0.00" if off < 3 else "#,##0"

    ws.column_dimensions["A"].width = 15
    for i in range(2, col_idx): ws.column_dimensions[get_column_letter(i)].width = 13
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf, len(payment_methods), len(branches)


def create_foodics_daily_avg_report(df, dates):
    if not dates: return None, None, 0
    num_days = len(dates)
    summary = df.groupby(["Payment Method"]).agg({
        "Net Amount": "sum", "Amount": "sum", "Return Amount": "sum", "Count": "sum"
    }).reset_index()
    summary["Avg Net Amount/Day"] = summary["Net Amount"] / num_days
    summary["Avg Count/Day"]      = summary["Count"] / num_days
    summary = summary.sort_values("Avg Net Amount/Day", ascending=False)

    wb = Workbook(); ws = wb.active; ws.title = "Daily_Averages"
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    avg_fill    = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    total_fill  = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")

    ws.cell(row=1, column=1, value=f"Report Period: {dates[0]} to {dates[-1]} ({num_days} days)")
    ws.cell(row=1, column=1).font = Font(bold=True, size=12)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)

    for col, header in enumerate(["Payment Method","Total Net Amount","Daily Avg Net",
                                   "Total Count","Daily Avg Count","Total Returns","Days"], 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.fill = header_fill; cell.font = Font(color="FFFFFF", bold=True, size=11); cell.alignment = center

    row_idx = 3
    for _, data in summary.iterrows():
        ws.cell(row=row_idx, column=1, value=data["Payment Method"])
        ws.cell(row=row_idx, column=2, value=data["Net Amount"])
        ws.cell(row=row_idx, column=3, value=data["Avg Net Amount/Day"])
        ws.cell(row=row_idx, column=4, value=data["Count"])
        ws.cell(row=row_idx, column=5, value=data["Avg Count/Day"])
        ws.cell(row=row_idx, column=6, value=data["Return Amount"])
        ws.cell(row=row_idx, column=7, value=num_days)
        ws.cell(row=row_idx, column=3).fill = avg_fill
        ws.cell(row=row_idx, column=5).fill = avg_fill
        for col in range(2, 7): ws.cell(row=row_idx, column=col).number_format = "#,##0.00" if col != 4 else "#,##0"
        row_idx += 1

    row_idx += 1
    for col, val in [(1,"GRAND TOTAL"),(2,summary["Net Amount"].sum()),
                     (3,summary["Net Amount"].sum()/num_days),(4,summary["Count"].sum()),
                     (5,summary["Count"].sum()/num_days),(6,summary["Return Amount"].sum()),(7,num_days)]:
        cell = ws.cell(row=row_idx, column=col, value=val)
        cell.fill = total_fill; cell.font = Font(bold=True, size=12)
        if col >= 2 and col != 4: cell.number_format = "#,##0.00"

    for ltr, w in zip("ABCDEFG", [30, 18, 18, 15, 18, 15, 10]): ws.column_dimensions[ltr].width = w
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf, summary, num_days




# ==================== ZENPUT FINANCIAL FORM ====================

ZENPUT_API_KEY       = "8051c8104fd221694d9aeb305f7f4abb"
ZENPUT_FINANCIAL_TID = 1556454    # Financial form template ID
ZENPUT_SHEET_ID      = "1QmcxxyEyDJTyaWp9zg5shcVHOLZPH7ibgr5STFT1ZhY"

# Zenput location code → branch label (same order/names as the rest of the app)
ZENPUT_BRANCH_MAP = {
    "2197299": "LBRUH B07",  "2239240": "FYJED B32",  "2235670": "ANRUH B31",
    "2190657": "SLAHS B23",  "2164026": "NDRUH B15",  "2164019": "SWRUH B08",
    "2203271": "SARUH B27",  "2164017": "DARUH B06",  "2164032": "KRRUH B21",
    "2164031": "SFJED B24",  "2164025": "RBRUH B14",  "2164016": "RWRUH B05",
    "2197297": "NSRUH B04",  "2164021": "SHRUH B10",  "2164013": "KHRUH B02",
    "2155652": "NURUH B01",  "2164023": "TWRUH B12",  "2164020": "AZRUH B09",
    "2199002": "RWAHS B25",  "2242934": "HIRJED B33", "2164022": "NRRUH B11",
    "2164030": "MURUH B19",  "2164014": "GHRUH B03",  "2211854": "QARUH B30",
    "2258220": "PSJED B36",  "2185452": "OBJED B22",  "2243963": "URRUH B34",
    "2199835": "HAJED B26",  "2210205": "MAJED B28",  "2250799": "IRRUH B35",
    "2164027": "BDRUH B16",  "2155654": "AQRUH B13",  "2197298": "TKRUH B18",
    "2164028": "QRRUH B17",  "2257790": "SHMAK B37",  "2260889": "UHDMM B38",
    "2263062": "HSRUH B39",
}

# The delivery/payment channels in the form (in order)
# Each channel has: Amount value  +  "Invoices - عدد الفواتير" sub-field
ZENPUT_CHANNELS = [
    "Noon - نون",
    "To you - تو يو",
    "Barakah - بركه",
    "Mr. Manddob - مستر مندوب",
    "Ninja - نينجا",
    "The chefz - ذا شيفز",
    "Marsool - مرسول",
    "Solo loyalty - سولو لوياليتي",
    "Jahez - جاهز",
    "Hungerstation - هنقرستيشن",
    "Ketta - كيتا",
    "Mada - مدى",
    "Cash - كاش",
    "Cash used without invoice - الكاش المستخدم من غير فاتورة",
    "Cash purchase invoice - فواتير الشراء النقدية",
]

# Short display names for columns
ZENPUT_CHANNEL_SHORT = {
    "Noon - نون":                                              "Noon",
    "To you - تو يو":                                         "To You",
    "Barakah - بركه":                                         "Barakah",
    "Mr. Manddob - مستر مندوب":                               "Mr. Manddob",
    "Ninja - نينجا":                                          "Ninja",
    "The chefz - ذا شيفز":                                    "The Chefz",
    "Marsool - مرسول":                                        "Marsool",
    "Solo loyalty - سولو لوياليتي":                           "Solo Loyalty",
    "Jahez - جاهز":                                           "Jahez",
    "Hungerstation - هنقرستيشن":                              "Hungerstation",
    "Ketta - كيتا":                                           "Ketta",
    "Mada - مدى":                                             "Mada",
    "Cash - كاش":                                             "Cash",
    "Cash used without invoice - الكاش المستخدم من غير فاتورة": "Cash With No Invoices",
    "Cash purchase invoice - فواتير الشراء النقدية":          "Cash Purchase Inv.",
}


def zenput_fetch_financial(start_date_str, end_date_str):
    """
    Fetch submissions for ZENPUT_FINANCIAL_TID between start and end dates.
    Returns a DataFrame with one row per submission, columns:
      Date | Branch | <channel> Amount | <channel> Invoices | ...
      Total Invoices | Total Sales | Foodics Sales | Difference | Notes
    Sorted by Date then by branch code order.
    """
    import requests, pytz
    from datetime import datetime

    TZ       = pytz.timezone("Asia/Baghdad")
    start_dt = TZ.localize(datetime.strptime(start_date_str, "%Y-%m-%d").replace(hour=0,  minute=0,  second=0))
    end_dt   = TZ.localize(datetime.strptime(end_date_str,   "%Y-%m-%d").replace(hour=23, minute=59, second=59))

    headers  = {"X-API-TOKEN": ZENPUT_API_KEY, "Content-Type": "application/json"}
    all_subs = []
    offset, limit = 0, 100

    while True:
        resp = requests.get(
            "https://www.zenput.com/api/v3/submissions/",
            headers=headers,
            params={
                "form_template_id": ZENPUT_FINANCIAL_TID,
                "limit":                limit,
                "start":                offset,
                "date_submitted_start": start_date_str,
            },
            timeout=30,
        )
        if resp.status_code != 200:
            raise Exception(f"Zenput API error {resp.status_code}: {resp.text[:300]}")

        batch = resp.json().get("data", [])
        if not batch:
            break

        for s in batch:
            raw_date = s.get("smetadata", {}).get("date_submitted_local", "")
            if not raw_date:
                continue
            try:
                sub_dt = datetime.fromisoformat(raw_date)
                if sub_dt.tzinfo is None:
                    sub_dt = TZ.localize(sub_dt)
                if start_dt <= sub_dt <= end_dt:
                    all_subs.append(s)
            except Exception:
                pass

        offset += limit

        # Early exit: if the oldest submission in this batch is before our range
        last_raw = batch[-1].get("smetadata", {}).get("date_submitted_local", "")
        if last_raw:
            try:
                last_plain = datetime.fromisoformat(last_raw).replace(tzinfo=None)
                if last_plain < datetime.strptime(start_date_str, "%Y-%m-%d"):
                    break
            except Exception:
                pass

    if not all_subs:
        return pd.DataFrame()

    # ── Parse each submission ──────────────────────────────────────────────
    rows = []
    for s in all_subs:
        raw_date = s.get("smetadata", {}).get("date_submitted_local", "")
        date_str = raw_date[:10] if raw_date else ""
        answers  = s.get("answers", [])

        row = {"Date": date_str, "Branch": ""}

        # Find branch from "Store - الفرع" field
        for ans in answers:
            t = ans.get("title", "")
            if "Store" in t or "الفرع" in t:
                code = str(ans.get("value", "")).strip()
                row["Branch"] = ZENPUT_BRANCH_MAP.get(code, code)
                break

        # Parse channels: each channel title appears once (amount),
        # followed immediately by "Invoices - عدد الفواتير" (count).
        i = 0
        while i < len(answers):
            ans   = answers[i]
            title = ans.get("title", "").strip()

            # Check channel match (partial match to handle minor title variations)
            matched = None
            for ch in ZENPUT_CHANNELS:
                if title == ch or title.startswith(ch[:20]):
                    matched = ch
                    break

            if matched:
                short      = ZENPUT_CHANNEL_SHORT[matched]
                amt        = ans.get("value", 0) or 0
                inv        = 0
                # Next answer should be Invoices sub-field
                if i + 1 < len(answers):
                    nxt = answers[i + 1]
                    if "Invoices" in nxt.get("title", "") or "فاتور" in nxt.get("title", ""):
                        inv = nxt.get("value", 0) or 0
                        i += 1
                row[f"{short} - Amount"]   = amt
                row[f"{short} - Invoices"] = inv

            elif any(k in title for k in ["Total invoices", "مجموع عدد الفواتير"]):
                row["Total Invoices"] = ans.get("value", 0) or 0

            elif any(k in title for k in ["Total cash & credit", "اجمالي المبيعات"]):
                row["Total Sales"] = ans.get("value", 0) or 0

            elif any(k in title for k in ["Sales by Foodics", "المبيعات في فوديكس"]):
                row["Foodics Sales"] = ans.get("value", 0) or 0

            elif any(k in title for k in ["Difference", "الفرق"]):
                row["Difference"] = ans.get("value", 0) or 0

            elif any(k in title for k in ["Notes", "ملاحظات"]):
                row["Notes"] = ans.get("value", "") or ""

            i += 1

        rows.append(row)

    df = pd.DataFrame(rows)

    # Sort: Date ascending, then branch by code order
    df["_sort"] = df["Branch"].apply(branch_code_sort_key)
    df = df.sort_values(["Date", "_sort"]).drop("_sort", axis=1).reset_index(drop=True)

    return df


def _zenput_group_df(df):
    """
    Group the raw Zenput DataFrame by Date + Branch, summing all numeric columns
    and joining Notes. Returns a cleaned, sorted DataFrame.
    Sheet structure:
      - One tab per date (named by date)
      - If only one date, single tab named by that date
      - Each tab: rows = branches, cols = channels
      - Bottom: TOTAL row per tab
    """
    if df.empty:
        return df

    # Separate text vs numeric cols
    text_cols    = ["Date", "Branch", "Notes"]
    numeric_cols = [c for c in df.columns if c not in text_cols]

    # Coerce all numeric cols to numbers
    df_work = df.copy()
    for c in numeric_cols:
        df_work[c] = pd.to_numeric(df_work[c], errors="coerce").fillna(0)

    # Group: sum numerics, join notes
    grp = df_work.groupby(["Date", "Branch"], sort=False)
    agg_dict = {c: "sum" for c in numeric_cols}
    if "Notes" in df_work.columns:
        agg_dict["Notes"] = lambda x: " | ".join(v for v in x if str(v).strip())
    grouped = grp.agg(agg_dict).reset_index()

    # Sort by date then branch code order
    grouped["_sort"] = grouped["Branch"].apply(branch_code_sort_key)
    grouped = grouped.sort_values(["Date", "_sort"]).drop("_sort", axis=1).reset_index(drop=True)

    # Drop: all Invoices columns EXCEPT Cash and Cash Purchase Inv.
    # Drop: Notes and Total Invoices entirely
    # Drop ALL Invoices columns and Notes and Total Invoices
    drop_cols = [
        c for c in grouped.columns
        if (c in ("Notes", "Total Invoices")) or ("Invoices" in c)
    ]
    grouped = grouped.drop(columns=drop_cols, errors="ignore")

    # Reorder cols: Date, Branch, channels..., summary cols
    priority_end = ["Total Sales", "Foodics Sales", "Difference"]
    middle = [c for c in grouped.columns if c not in ["Date", "Branch"] + priority_end]
    final_cols = ["Date", "Branch"] + middle + [c for c in priority_end if c in grouped.columns]
    return grouped[final_cols]


def zenput_build_excel(df):
    """
    Build a formatted Excel workbook from the Zenput financial DataFrame.
    - Groups by Date + Branch (sums all numeric cols)
    - One sheet per date, named by the date (e.g. '11-Mar')
    - Each sheet: rows = branches, cols = channels, bottom = TOTAL row
    Returns a BytesIO buffer.
    """
    df_grouped = _zenput_group_df(df)
    if df_grouped.empty:
        buf = io.BytesIO()
        Workbook().save(buf)
        buf.seek(0)
        return buf

    wb = Workbook()
    wb.remove(wb.active)   # remove default sheet

    header_fill  = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    branch_fill  = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
    amount_fill  = PatternFill(start_color="DEEAF1", end_color="DEEAF1", fill_type="solid")
    invoice_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    total_fill   = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    summary_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    grand_fill   = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right  = Alignment(horizontal="right")
    left   = Alignment(horizontal="left")

    dates = sorted(df_grouped["Date"].unique())
    cols  = [c for c in df_grouped.columns if c != "Date"]   # Branch + all data cols

    # Column type classification
    amount_cols  = [c for c in cols if "Amount"  in c or "Sales"  in c or "Difference" in c]
    invoice_cols = [c for c in cols if "Invoices" in c and c != "Total Invoices"]
    total_inv    = [c for c in cols if c == "Total Invoices"]
    total_sales  = [c for c in cols if c in ("Total Sales", "Foodics Sales")]
    notes_cols   = [c for c in cols if c == "Notes"]

    for date_str in dates:
        # Tab name: short date like "11-Mar"
        try:
            import datetime as _dti
            tab_name = _dti.datetime.strptime(date_str, "%Y-%m-%d").strftime("%d-%b")
        except Exception:
            tab_name = str(date_str)[:10]

        ws = wb.create_sheet(title=tab_name)
        df_date = df_grouped[df_grouped["Date"] == date_str].reset_index(drop=True)

        # ── Date banner row ──────────────────────────────────────────────────
        date_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        ws.cell(row=1, column=1, value=f"Financial Summary — {tab_name}").fill = date_fill
        ws.cell(row=1, column=1).font      = Font(color="FFFFFF", bold=True, size=12)
        ws.cell(row=1, column=1).alignment = center
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))

        # ── Column headers ───────────────────────────────────────────────────
        for c_idx, col_name in enumerate(cols, 1):
            cell = ws.cell(row=2, column=c_idx, value=col_name)
            cell.font      = Font(color="FFFFFF", bold=True, size=9)
            cell.alignment = center
            # Colour-code header by column type
            if col_name == "Branch":
                cell.fill = branch_fill
            elif col_name in amount_cols:
                cell.fill = PatternFill(start_color="1A5276", end_color="1A5276", fill_type="solid")
            elif col_name in invoice_cols:
                cell.fill = PatternFill(start_color="1E8449", end_color="1E8449", fill_type="solid")
            elif col_name in total_inv + total_sales:
                cell.fill = PatternFill(start_color="B7770D", end_color="B7770D", fill_type="solid")
            elif col_name == "Difference":
                cell.fill = PatternFill(start_color="922B21", end_color="922B21", fill_type="solid")
            elif col_name == "Notes":
                cell.fill = PatternFill(start_color="566573", end_color="566573", fill_type="solid")
            else:
                cell.fill = header_fill

        # ── Data rows ────────────────────────────────────────────────────────
        for r_idx, row_data in enumerate(df_date.itertuples(index=False), 3):
            row_dict = dict(zip(df_date.columns, row_data))
            for c_idx, col_name in enumerate(cols, 1):
                val  = row_dict.get(col_name, "")
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                if col_name == "Branch":
                    cell.font      = Font(bold=True, size=9)
                    cell.alignment = left
                elif col_name in amount_cols:
                    cell.fill         = amount_fill
                    cell.number_format = "#,##0.00"
                    cell.alignment    = right
                    cell.font         = Font(size=9)
                elif col_name in invoice_cols:
                    cell.fill         = invoice_fill
                    cell.number_format = "#,##0"
                    cell.alignment    = right
                    cell.font         = Font(size=9)
                elif col_name in total_inv:
                    cell.fill         = total_fill
                    cell.font         = Font(bold=True, size=9)
                    cell.number_format = "#,##0"
                    cell.alignment    = right
                elif col_name in total_sales:
                    cell.fill         = total_fill
                    cell.font         = Font(bold=True, size=9)
                    cell.number_format = "#,##0.00"
                    cell.alignment    = right
                elif col_name == "Difference":
                    cell.fill         = summary_fill
                    cell.font         = Font(bold=True, size=9)
                    cell.number_format = "#,##0.00"
                    cell.alignment    = right
                elif col_name == "Notes":
                    cell.alignment = left
                    cell.font      = Font(size=9)

        # ── TOTAL row ────────────────────────────────────────────────────────
        total_row_idx = 3 + len(df_date)
        numeric_sum_cols = [c for c in cols if c not in ("Branch", "Notes")]
        for c_idx, col_name in enumerate(cols, 1):
            cell = ws.cell(row=total_row_idx, column=c_idx)
            cell.fill = grand_fill
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            if col_name == "Branch":
                cell.value     = "TOTAL"
                cell.alignment = center
            elif col_name in numeric_sum_cols:
                col_vals = pd.to_numeric(df_date[col_name], errors="coerce").fillna(0)
                cell.value = col_vals.sum()
                if col_name in invoice_cols + total_inv:
                    cell.number_format = "#,##0"
                else:
                    cell.number_format = "#,##0.00"
                cell.alignment = right
            else:
                cell.value = ""

        # ── Column widths ─────────────────────────────────────────────────────
        ws.column_dimensions["A"].width = 16   # Branch
        for i in range(2, len(cols) + 1):
            col_name = cols[i - 1]
            if col_name == "Notes":
                ws.column_dimensions[get_column_letter(i)].width = 25
            elif "Invoices" in col_name:
                ws.column_dimensions[get_column_letter(i)].width = 9
            else:
                ws.column_dimensions[get_column_letter(i)].width = 14

        # Freeze header rows
        ws.freeze_panes = "B3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def zenput_build_excel_by_branch(df):
    """
    Build an Excel workbook grouped purely by Branch (all dates combined, summed).
    Single sheet — one row per branch, TOTAL row at bottom.
    """
    df_grouped = _zenput_group_df(df)
    if df_grouped.empty:
        buf = io.BytesIO()
        Workbook().save(buf)
        buf.seek(0)
        return buf

    # Drop Date, group by Branch only
    cols_no_date = [c for c in df_grouped.columns if c != "Date"]
    numeric_cols = [c for c in cols_no_date if c not in ("Branch", "Notes")]

    df_work = df_grouped[cols_no_date].copy()
    for c in numeric_cols:
        df_work[c] = pd.to_numeric(df_work[c], errors="coerce").fillna(0)

    agg_dict = {c: "sum" for c in numeric_cols}
    if "Notes" in df_work.columns:
        agg_dict["Notes"] = lambda x: " | ".join(v for v in x if str(v).strip())

    df_branch = df_work.groupby("Branch", sort=False).agg(agg_dict).reset_index()
    df_branch["_sort"] = df_branch["Branch"].apply(branch_code_sort_key)
    df_branch = df_branch.sort_values("_sort").drop("_sort", axis=1).reset_index(drop=True)

    # Reorder cols
    priority_end = ["Total Invoices", "Total Sales", "Foodics Sales", "Difference", "Notes"]
    middle = [c for c in df_branch.columns if c not in ["Branch"] + priority_end]
    final_cols = ["Branch"] + middle + [c for c in priority_end if c in df_branch.columns]
    df_branch = df_branch[final_cols]

    cols = df_branch.columns.tolist()
    amount_cols  = [c for c in cols if "Amount"   in c or "Sales"      in c or "Difference" in c]
    invoice_cols = [c for c in cols if "Invoices" in c and c != "Total Invoices"]
    total_inv    = [c for c in cols if c == "Total Invoices"]
    total_sales  = [c for c in cols if c in ("Total Sales", "Foodics Sales")]

    wb = Workbook()
    ws = wb.active
    ws.title = "By Branch"

    header_fill  = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    amount_fill  = PatternFill(start_color="DEEAF1", end_color="DEEAF1", fill_type="solid")
    invoice_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    total_fill   = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    summary_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    grand_fill   = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right  = Alignment(horizontal="right")
    left   = Alignment(horizontal="left")

    # Header row
    for c_idx, col_name in enumerate(cols, 1):
        cell = ws.cell(row=1, column=c_idx, value=col_name)
        cell.font      = Font(color="FFFFFF", bold=True, size=9)
        cell.alignment = center
        if col_name == "Branch":
            cell.fill = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
        elif col_name in amount_cols:
            cell.fill = PatternFill(start_color="1A5276", end_color="1A5276", fill_type="solid")
        elif col_name in invoice_cols:
            cell.fill = PatternFill(start_color="1E8449", end_color="1E8449", fill_type="solid")
        elif col_name in total_inv + total_sales:
            cell.fill = PatternFill(start_color="B7770D", end_color="B7770D", fill_type="solid")
        elif col_name == "Difference":
            cell.fill = PatternFill(start_color="922B21", end_color="922B21", fill_type="solid")
        elif col_name == "Notes":
            cell.fill = PatternFill(start_color="566573", end_color="566573", fill_type="solid")
        else:
            cell.fill = header_fill

    # Data rows
    for r_idx, row_data in enumerate(df_branch.itertuples(index=False), 2):
        row_dict = dict(zip(cols, row_data))
        for c_idx, col_name in enumerate(cols, 1):
            val  = row_dict.get(col_name, "")
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if col_name == "Branch":
                cell.font = Font(bold=True, size=9); cell.alignment = left
            elif col_name in amount_cols:
                cell.fill = amount_fill; cell.number_format = "#,##0.00"
                cell.alignment = right; cell.font = Font(size=9)
            elif col_name in invoice_cols:
                cell.fill = invoice_fill; cell.number_format = "#,##0"
                cell.alignment = right; cell.font = Font(size=9)
            elif col_name in total_inv:
                cell.fill = total_fill; cell.font = Font(bold=True, size=9)
                cell.number_format = "#,##0"; cell.alignment = right
            elif col_name in total_sales:
                cell.fill = total_fill; cell.font = Font(bold=True, size=9)
                cell.number_format = "#,##0.00"; cell.alignment = right
            elif col_name == "Difference":
                cell.fill = summary_fill; cell.font = Font(bold=True, size=9)
                cell.number_format = "#,##0.00"; cell.alignment = right
            elif col_name == "Notes":
                cell.alignment = left; cell.font = Font(size=9)

    # TOTAL row
    total_row_idx = 2 + len(df_branch)
    numeric_sum_cols = [c for c in cols if c not in ("Branch", "Notes")]
    for c_idx, col_name in enumerate(cols, 1):
        cell = ws.cell(row=total_row_idx, column=c_idx)
        cell.fill = grand_fill; cell.font = Font(bold=True, color="FFFFFF", size=10)
        if col_name == "Branch":
            cell.value = "TOTAL"; cell.alignment = center
        elif col_name in numeric_sum_cols:
            cell.value = pd.to_numeric(df_branch[col_name], errors="coerce").fillna(0).sum()
            cell.number_format = "#,##0" if col_name in invoice_cols + total_inv else "#,##0.00"
            cell.alignment = right
        else:
            cell.value = ""

    # Column widths
    ws.column_dimensions["A"].width = 16
    for i in range(2, len(cols) + 1):
        col_name = cols[i - 1]
        if col_name == "Notes":
            ws.column_dimensions[get_column_letter(i)].width = 25
        elif "Invoices" in col_name:
            ws.column_dimensions[get_column_letter(i)].width = 9
        else:
            ws.column_dimensions[get_column_letter(i)].width = 14

    ws.freeze_panes = "B2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def zenput_push_to_sheet(df, tab_name):
    """Push the financial DataFrame to a tab in ZENPUT_SHEET_ID with formatting."""
    try:
        gc = get_gspread_client()
    except Exception as e:
        return False, f"Auth failed: {str(e)}"
    try:
        sh = gc.open_by_key(ZENPUT_SHEET_ID)
    except Exception as e:
        return False, f"Cannot open sheet — share it with the service account first.\nError: {str(e)}"
    try:
        try:
            ws = sh.worksheet(tab_name)
            ws.clear()
            if ws.row_count < len(df) + 10 or ws.col_count < len(df.columns) + 5:
                ws.resize(rows=len(df) + 10, cols=len(df.columns) + 5)
        except Exception:
            ws = sh.add_worksheet(title=tab_name, rows=len(df) + 10, cols=len(df.columns) + 5)

        rows = [df.columns.tolist()] + df.fillna("").astype(str).values.tolist()
        ws.update(rows, value_input_option="USER_ENTERED")

        # Formatting: header row + freeze + auto-resize
        def _rgb(r, g, b):
            return {"red": r/255, "green": g/255, "blue": b/255}

        n_cols = len(df.columns)
        sh.batch_update({"requests": [
            # Header: dark blue bg, white bold text
            {"repeatCell": {
                "range": {"sheetId": ws.id, "startRowIndex": 0, "endRowIndex": 1,
                          "startColumnIndex": 0, "endColumnIndex": n_cols},
                "cell": {"userEnteredFormat": {
                    "backgroundColor": _rgb(31, 78, 120),
                    "textFormat": {"bold": True, "fontSize": 9,
                                   "foregroundColor": {"red": 1, "green": 1, "blue": 1}},
                    "horizontalAlignment": "CENTER",
                    "wrapStrategy": "WRAP",
                }},
                "fields": "userEnteredFormat(backgroundColor,textFormat,horizontalAlignment,wrapStrategy)"
            }},
            # Freeze header row
            {"updateSheetProperties": {
                "properties": {"sheetId": ws.id, "gridProperties": {"frozenRowCount": 1}},
                "fields": "gridProperties.frozenRowCount"
            }},
            # Auto-resize all columns
            {"autoResizeDimensions": {
                "dimensions": {"sheetId": ws.id, "dimension": "COLUMNS",
                               "startIndex": 0, "endIndex": n_cols}
            }},
        ]})
        return True, f"Pushed {len(df)} rows to tab '{tab_name}'"
    except Exception as e:
        import traceback
        return False, f"Write error: {str(e)}\n{traceback.format_exc()}"


# ==================== UI ====================

st.title("🏦 Geidea & Foodics Summary Generator")
st.markdown("Upload your reconciliation file to generate summary reports")

uploaded_file = st.file_uploader(
    "📁 Upload file (Geidea or Foodics) — supports .xlsx, .xls, .csv",
    type=["xlsx", "xls", "csv"]
)

if uploaded_file:
    try:
        df_raw    = read_uploaded_file(uploaded_file)
        file_type = detect_file_type(df_raw)

        # ── GEIDEA ──────────────────────────────────────────────────────────────
        if file_type == "geidea":
            st.success(f"✅ Detected **Geidea** file: {uploaded_file.name} ({len(df_raw)} rows)")

            with st.expander("🔍 Preview Raw Data (with row numbers)"):
                preview_df = df_raw.copy()
                preview_df.index = range(1, len(preview_df) + 1)   # 1-based for user
                st.dataframe(preview_df, use_container_width=True, height=300)

            # ── Previous Day row selection ─────────────────────────────────────
            st.markdown("---")
            st.subheader("📅 Previous Day Rows (Optional)")
            st.markdown(
                "The raw file includes a **timestamp** column. Rows from a previous day "
                "(before midnight cutoff) can appear mixed in. Select the row range below "
                "to split them into a separate **'Previous Day'** sheet in the simplified report."
            )
            total_rows = len(df_raw)
            use_prev = st.checkbox("✂️ Split previous day rows into a separate sheet", value=False)

            df_prev_raw = None
            df_today_raw = df_raw.copy()

            if use_prev:
                col_a, col_b = st.columns(2)
                with col_a:
                    prev_start = st.number_input(
                        "Previous day: first row", min_value=1, max_value=total_rows,
                        value=1, step=1,
                        help="Row number in the file (1 = first data row)"
                    )
                with col_b:
                    prev_end = st.number_input(
                        "Previous day: last row", min_value=1, max_value=total_rows,
                        value=min(10, total_rows), step=1
                    )
                if prev_start > prev_end:
                    st.error("⚠️ Start row must be ≤ end row.")
                    use_prev = False
                else:
                    # Rows are 1-based in UI, 0-based in pandas
                    prev_idx   = list(range(prev_start - 1, prev_end))
                    today_idx  = [i for i in range(total_rows) if i not in set(prev_idx)]
                    df_prev_raw  = df_raw.iloc[prev_idx].reset_index(drop=True)
                    df_today_raw = df_raw.iloc[today_idx].reset_index(drop=True)
                    # Detect the base reconciliation date and compute labels
                    import datetime as _dt_ui
                    _date_col = next((c for c in df_raw.columns if "date" in c.lower() and "recon" in c.lower()), None)
                    _today_date_str = "today"
                    _prev_date_str  = "previous day"
                    if _date_col:
                        try:
                            _base_date = pd.to_datetime(df_today_raw[_date_col], errors="coerce").dt.date.dropna().unique()
                            if len(_base_date):
                                _d = sorted(_base_date)[-1]
                                _today_date_str = _d.strftime("%d-%b")
                                _prev_date_str  = (_d - _dt_ui.timedelta(days=1)).strftime("%d-%b")
                        except Exception:
                            pass
                    st.info(
                        f"📌 **Rows {prev_start}–{prev_end}** → sheet **`{_prev_date_str}`** ({len(prev_idx)} rows)  |  "
                        f"**Remaining** → sheet **`{_today_date_str}`** ({len(today_idx)} rows)"
                    )
                    with st.expander("👁 Preview Previous Day rows"):
                        st.dataframe(df_prev_raw, use_container_width=True, height=200)

            st.markdown("---")

            with st.spinner("Processing Geidea reports..."):
                df_processed       = process_geidea_data(df_today_raw)
                df_prev_processed  = process_geidea_data(df_prev_raw) if df_prev_raw is not None else None

                summary_buffer, summary_df, grand_total = create_geidea_summary_file(df_processed)
                detailed_buffer, num_terminals          = create_geidea_detailed_file(df_processed)
                detailed_tot_buffer, _                  = create_geidea_detailed_totals_only(
                                                              df_processed, df_prev_processed)

                unique_dates     = df_processed["Reconciliation Date"].dropna().unique()
                has_multiple_dates = len(unique_dates) > 1
                if has_multiple_dates:
                    summary_date_buffer, _, num_dates = create_geidea_summary_by_date_file(df_processed)
                    date_buffer, _, _                 = create_geidea_detailed_by_date_file(df_processed)
                else:
                    summary_date_buffer = date_buffer = None; num_dates = 0

            st.subheader("📊 Geidea Summary Preview")
            col1, col2, col3 = st.columns(3)
            col1.metric("Banks",        summary_df["Bank Name"].nunique())
            col2.metric("Card Schemes", summary_df["Card Name"].nunique())
            col3.metric("Grand Total",  f"{grand_total:,.0f}")

            if has_multiple_dates:
                st.info(f"📅 Detected {num_dates} dates: {', '.join([d.strftime('%Y-%m-%d') for d in unique_dates])}")
            if "Unknown Bank" in summary_df["Bank Name"].values:
                st.warning("⚠️ Some terminals not found in mapping (shown in red)")

            st.dataframe(
                summary_df.style.format({"Total": "{:,.2f}"})
                .apply(lambda x: ["background-color: #FF6B6B; color: white"] * 3
                       if x["Bank Name"] == "Unknown Bank" else [""] * 3, axis=1),
                use_container_width=True, height=300
            )

            st.subheader("⬇️ Geidea: Download Reports")
            n_reports = 5 if has_multiple_dates else 3
            prev_note = " *(includes Previous Day sheet)*" if use_prev and df_prev_raw is not None else ""
            st.markdown(f"**{n_reports} Geidea reports generated:**")
            c1, c2 = st.columns(2)
            with c1:
                st.download_button("📊 Summary Totals Only", data=summary_buffer,
                    file_name="Geidea_01_SUMMARY_Totals_Only.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True)
                st.download_button("📋 Detailed — Full (Debit / Credit / Total)", data=detailed_buffer,
                    file_name=f"Geidea_02_DETAILED_Full_{num_terminals}_terminals.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True)
                st.download_button(
                    f"📋 Detailed — Total per Terminal{prev_note}",
                    data=detailed_tot_buffer,
                    file_name=f"Geidea_03_DETAILED_Total_{num_terminals}_terminals.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True)
            with c2:
                if has_multiple_dates:
                    st.download_button("📅 Summary by Date", data=summary_date_buffer,
                        file_name=f"Geidea_04_SUMMARY_by_Date_{num_dates}_dates.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True)
                    st.download_button("📆 Detailed by Date (Full)", data=date_buffer,
                        file_name=f"Geidea_05_DETAILED_by_Date_{num_dates}_dates.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True)
            st.success(f"✅ All {n_reports} Geidea reports ready! ({num_terminals} terminals)")

        # ── FOODICS ─────────────────────────────────────────────────────────────
        elif file_type == "foodics":
            st.success(f"✅ Detected **Foodics** Payments Report: {uploaded_file.name}")
            with st.expander("🔍 Preview Raw Data"):
                st.dataframe(df_raw.head(15), use_container_width=True)

            with st.spinner("Processing Foodics reports..."):
                df_processed, dates = process_foodics_data(df_raw)

                branch_buffer,   branch_summary, num_branches = create_foodics_summary_by_branch(df_processed)
                payment_buffer,  payment_summary              = create_foodics_summary_by_payment_method(df_processed)
                det_br_full_buf, num_br_det, _                = create_foodics_detailed_by_branch(df_processed)
                det_pm_full_buf, num_pm_cols, _               = create_foodics_detailed_by_payment_method(df_processed)

                det_br_net_buf, _, _ = _foodics_net_only_pivot(
                    df_processed, row_field="Payment Method", col_field="Branch",
                    row_label="Payment Method", col_label="Branch",
                    header_hex="2E7D32", sub_hex="A5D6A7"
                )
                det_pm_net_buf, _, _ = _foodics_net_only_pivot(
                    df_processed, row_field="Branch", col_field="Payment Method",
                    row_label="Branch", col_label="Payment Method",
                    header_hex="1565C0", sub_hex="90CAF9"
                )

                if dates:
                    avg_buffer, _, num_days = create_foodics_daily_avg_report(df_processed, dates)
                else:
                    avg_buffer = None; num_days = 0

            st.subheader("📊 Foodics Summary Preview")
            col1, col2, col3 = st.columns(3)
            col1.metric("Branches",        num_branches)
            col2.metric("Payment Methods", payment_summary["Payment Method"].nunique())
            col3.metric("Total Net Amount",f"{payment_summary['Net Amount'].sum():,.0f}")

            if dates:
                st.info(f"📅 Report period: {dates[0]} to {dates[-1]} ({num_days} days)")
            else:
                st.info("ℹ️ No date range metadata — daily averages not available for plain CSV.")

            st.dataframe(
                payment_summary.style.format({
                    "Net Amount": "{:,.2f}", "Amount": "{:,.2f}",
                    "Return Amount": "{:,.2f}", "Count": "{:,.0f}"
                }),
                use_container_width=True, height=300
            )

            st.subheader("⬇️ Foodics: Download Reports")
            total_reports = 7 if dates else 6
            st.markdown(f"**{total_reports} Foodics reports generated:**")

            c1, c2 = st.columns(2)
            with c1:
                st.download_button("🏪 Summary by Branch", data=branch_buffer,
                    file_name=f"Foodics_01_SUMMARY_by_Branch_{num_branches}_branches.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True)
                st.download_button("📋 Detailed by Branch — Full (Net/Amount/Returns/Count)",
                    data=det_br_full_buf,
                    file_name=f"Foodics_03_DETAILED_Branch_Full_{num_br_det}_branches.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True)
                st.download_button("📋 Detailed by Branch — Net Total per Branch",
                    data=det_br_net_buf,
                    file_name=f"Foodics_05_DETAILED_Branch_Net_{num_br_det}_branches.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True)
                if dates:
                    st.download_button("📈 Daily Averages", data=avg_buffer,
                        file_name=f"Foodics_07_Daily_Averages_{num_days}_days.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True)
            with c2:
                st.download_button("💳 Summary by Payment Method", data=payment_buffer,
                    file_name="Foodics_02_SUMMARY_by_Payment_Method.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True)
                st.download_button("📊 Detailed by Payment Method — Full (Net/Amount/Returns/Count)",
                    data=det_pm_full_buf,
                    file_name=f"Foodics_04_DETAILED_PayMethod_Full_{num_pm_cols}_methods.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True)
                st.download_button("📊 Detailed by Payment Method — Net Total per Method",
                    data=det_pm_net_buf,
                    file_name=f"Foodics_06_DETAILED_PayMethod_Net_{num_pm_cols}_methods.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True)

            st.success(f"✅ {total_reports} Foodics reports ready! ({num_branches} branches · {payment_summary['Payment Method'].nunique()} payment methods)")

        else:
            st.error("❌ Could not detect file type.")
            st.info("**Geidea:** File must have 'Terminal' and 'Card Name' columns")
            st.info("**Foodics:** File must have 'Payment Method' and 'Branch' columns")

    except Exception as e:
        st.error(f"❌ Error processing file: {str(e)}")
        st.info("Please check your file format and try again.")

st.markdown("---")
st.header("📋 Zenput Financial Form")
st.markdown(
    "Fetch financial form submissions (template **1556454**) from Zenput, "
    "generate a formatted Excel summary, and optionally push to Google Sheets."
)

import datetime as _zdt
_today = _zdt.date.today()

z_c1, z_c2 = st.columns(2)
with z_c1:
    z_start = st.date_input("📅 From date", value=_today, key="z_start")
with z_c2:
    z_end = st.date_input("📅 To date", value=_today, key="z_end")

z_c3, z_c4 = st.columns(2)
with z_c3:
    z_dl   = st.checkbox("⬇️ Download as Excel", value=True,  key="z_dl")
with z_c4:
    z_push = st.checkbox("📤 Push to Google Sheet", value=False, key="z_push")

if z_push:
    z_tab_name = st.text_input(
        "Sheet tab name",
        value=f"Financial_{z_start.strftime('%d-%b')}" + (f"_to_{z_end.strftime('%d-%b')}" if z_end != z_start else ""),
        key="z_tab"
    )

if st.button("🚀 Fetch Zenput Submissions", type="primary", use_container_width=True, key="z_run"):
    if z_start > z_end:
        st.error("Start date must be ≤ end date.")
    else:
        with st.spinner(f"Fetching Zenput submissions {z_start} → {z_end} ..."):
            try:
                z_df = zenput_fetch_financial(str(z_start), str(z_end))
            except Exception as e:
                st.error(f"❌ Fetch error: {str(e)}")
                z_df = pd.DataFrame()

        if z_df.empty:
            st.warning("⚠️ No submissions found for the selected date range.")
        else:
            st.success(f"✅ Fetched **{len(z_df)} submissions** ({z_start} → {z_end})")
            with st.expander("👁 Preview data"):
                st.dataframe(z_df, use_container_width=True, height=300)

            if z_dl:
                _fname_base = f"Zenput_Financial_{z_start}{'_to_'+str(z_end) if z_end != z_start else ''}"
                z_buf_by_date   = zenput_build_excel(z_df)
                z_buf_by_branch = zenput_build_excel_by_branch(z_df)
                dl_c1, dl_c2 = st.columns(2)
                with dl_c1:
                    st.download_button(
                        "⬇️ By Date (one sheet per day)",
                        data=z_buf_by_date,
                        file_name=f"{_fname_base}_ByDate.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        key="z_dl_date"
                    )
                with dl_c2:
                    st.download_button(
                        "⬇️ By Branch (all dates combined)",
                        data=z_buf_by_branch,
                        file_name=f"{_fname_base}_ByBranch.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        key="z_dl_branch"
                    )

            if z_push:
                with st.spinner("Pushing to Google Sheets..."):
                    ok, msg = zenput_push_to_sheet(z_df, z_tab_name)
                if ok:
                    st.success(f"✅ {msg}")
                    st.markdown(f"[🔗 Open Google Sheet](https://docs.google.com/spreadsheets/d/{ZENPUT_SHEET_ID}/edit)")
                else:
                    st.error(msg)

st.markdown("---")
st.caption("Geidea & Foodics v5.8 | Branch-grouped simplified · Previous Day split · Zenput Financial")
